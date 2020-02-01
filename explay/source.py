# -*- coding: utf-8 -*-

from __future__ import absolute_import
from __future__ import division
from __future__ import print_function

import os
import re
import sys
import json
import glob
import yaml
import regex
import __main__
from copy import copy
import datetime
from collections import namedtuple, defaultdict
import inspect
import functools
import glob
import numpy as np

import xlrd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
#
from pandas import DataFrame
import pandas as pd
import pdb
#
from explay.utils import is_buildin, replace_str, to_yml
from explay.openpyxl_ext import insert_rows
from explay.parser import xlParser, xlBinaryParser
from explay.merger import xlMerger, xlConverter

from explay.utils import register_func
from explay.agg_func import agg_functions
from explay.post_func import common_funcs



def compose(*functions):
    def compose2(f, g):
        return lambda x: f(g(x))
    return functools.reduce(compose2, functions, lambda x: x)


class xlRenderer():
    def __init__(self, params):
        self.params = params
        self.first_row, self.idx_colname = params['first_row'], params['idx_colname']

    def __repr__(self):
        msg = '[Renderer]\n'
        msg += yaml.dump(self.params, indent=True, 
                allow_unicode=True, default_flow_style=False)
        return msg

    @classmethod
    def to_excel(cls, df, path):
        writer = pd.ExcelWriter(path, engine='xlsxwriter')
        df.to_excel(writer, sheet_name='Sheet1', index=False)

        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        datetime_format = workbook.add_format({'num_format': "yyyy-mm-dd"})
        time_format = workbook.add_format({'num_format': "hh:mm"})

        for index, values in df.iterrows():
            for i, v in enumerate(values):
                if type(v) == pd.Timestamp:
                    worksheet.write_datetime('%s%s' % (chr(65+i), index + 2), v, datetime_format)
                elif type(v) == datetime.time:
                    cell = '%s%s' % (chr(65+i), index+2)
                    worksheet.write_datetime('%s%s' % (chr(65+i), index + 2), v, time_format)
        writer.save()

    def render_excel(self, df, saved_name, template_name, template_dir='xls_template'):

        #  xls_file, xlsx_file = ['template/%s.%s' % (template_name, ext) for ext in ['xls', 'xlsx']]
        xls_file, xlsx_file = ['%s/%s.%s' % (template_dir, template_name, ext) for ext in ['xls', 'xlsx']]
        template = load_workbook(xlsx_file)

        rows = dataframe_to_rows(df, index=True, header=False)
        ws = template.active
        ws.insert_rows = insert_rows

        ws.insert_rows(ws, self.first_row, len(df))
        xls_from = xlrd.open_workbook(xls_file, formatting_info=True)
        xls_sheet = xls_from.sheet_by_index(0)
        color_index = lambda row, col: (
            xls_from.xf_list[xls_sheet.cell_xf_index(row,col)].background.pattern_colour_index)
        rgb = lambda row, col: xls_from.colour_map[color_index(row,col)]
        to_hexcode = lambda rgb_code: '00%s' % ''.join([('%x'% e).upper() for e in rgb_code])

        get_cell = lambda sht, row, col: sht['%s%d' % (get_column_letter(col), row)]

        for r_idx, row in enumerate(rows, self.first_row):
            for c_idx, df_colname in self.idx_colname.items():
                col_idx = df.columns.get_loc(df_colname)
                get_cell(ws, r_idx, c_idx).value = row[col_idx+1]

        for row in ws.rows:
            max_contents_len = []
            for cell in row:
                r_idx, c_idx = cell.row, cell.col_idx

                max_content_len = len(str(get_cell(ws, r_idx, c_idx)))
                max_contents_len.append(max_content_len)

                #  if cell.has_style:
                new_cell = get_cell(ws, r_idx, c_idx)

                if r_idx >= self.first_row:
                    font = get_cell(ws, self.first_row, c_idx).font
                else:
                    font = cell.font
                new_cell.font = copy(font)
                new_cell.border = copy(cell.border)

                # uses xlrd to get cell styles
                if r_idx < self.first_row:
                    color = to_hexcode(rgb(r_idx-1, c_idx-1))
                else:
                    color = to_hexcode(rgb(self.first_row-1, c_idx-1))
                fill = PatternFill(fill_type='solid',start_color=color)
                new_cell.fill = copy(fill)

                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

        template.save(saved_name)
        print('{} saved'.format(saved_name))


class xlTemplate():
    def __init__(self, params):
        self.params = params
        self._template = params['template']
        self.output = params['output']

    def __getitem__(self, index):
        return self._template[index]

    def __len__(self):
        return len(self._template)


class xlManager():

    def __init__(self, yml_file, home=None, local_imported=None):
        home = os.getcwd() if not home else home
        yml_path = '%s/%s' % (home, yml_file)
        self.yml, self.home = yml_path, home
        self.sources = dict()
        self.content = _c = to_yml(yml_file, True)
        self.converter = _c['xlconverter']
        self.parser = _c['xlparser']
        self.renderer = _c['xlrenderer']
        self.renderer = xlRenderer(self.renderer) if self.renderer else None

        self.load_parser()
        if local_imported:
            self.import_local(local_imported)

    def load_parser(self):
        parsers = {}
        for each in self.parser:
            print('\n')
            name = each['name']
            each_parser = None
            each = defaultdict(str, each)
            each_parser = xlBinaryParser(each)
            parsers[name] = each_parser
        self.parsers = parsers

    def import_local(self, local):
        for name, obj in self.parsers.items():
            local[name] = obj

    def import_source_df(self, local):
        for name, obj in self.parsers.items():
            local[name] = obj

    def import_variables(self, local):
        for v, value in self.variables.items():
            local[v] = value

    def register_func(self):
        import func
        #  print('reigister_func')
        funcs = [f for f in dir(func) if f.startswith('exp')]
        for func_name in funcs:
            func_name_in_yml = func_name[4:]
            #  print('func {} registed.'.format(func_name))
            #  print('func name in yml: {}'.format(func_name_in_yml))
            register_custom_func(func_name_in_yml, getattr(func, func_name))

    def load_excel(self, converter_name, filepath, sheet_name=0):
        cv = xlConverter(self.process.converter)
        df = cv.load_excel(converter_name, filepath, sheet_name)
        return df

    def merge_from_config(self, merger_yml_file, local_imported=None):
        merger_params = to_yml(merger_yml_file, True)['xlmerger']
        merged_all = []
        for each in merger_params:
            each = defaultdict(str, each)
            merge_type, name, location = each['type'], each['name'], each['location']
            print('\n', merge_type, name, location)
            converter_name = each['converter_name']
            if merge_type == 'merge_sheets':
                df_merged = self.merge_sheets(converter_name, location, each['sheet_names'])
            elif merge_type == 'merge_all':
                df_merged, _ = self.merge_all(converter_name, each['sheet_name'])
            elif merge_type == 'merge_files':
                df_merged = self.merge_files(converter_name, location, each['sheet_name'])
            merged_all.append(df_merged)
            if local_imported:
                local_imported[name] = df_merged
        return merged_all

    def merge_sheets(self, converter_name, filepath, sheet_names):
        source_path = os.path.join(self.home, 'source')
        self.merger = xlMerger(self.converter, source_path)
        df_merged = self.merger.merge_sheets(converter_name, filepath, sheet_names)
        self.sources[converter_name] = df_merged
        return df_merged

    def merge_files(self, converter_name, filepaths, sheet_name=0):
        source_path = os.path.join(self.home, 'source')
        self.merger = xlMerger(self.converter, source_path)
        df_merged = self.merger.merge_files(converter_name, filepaths, sheet_name)
        self.sources[converter_name] = df_merged
        return df_merged

    def merge_all(self, converter_name, sheet_name=0, filename_excludes=None, save=True):
        print(sheet_name, filename_excludes)
        source_path = os.path.join(self.home, 'source')
        self.merger = xlMerger(self.converter, source_path)
        df_merged, file_names = self.merger.merge_all(converter_name, sheet_name, filename_excludes)
        self.sources[converter_name] = df_merged
        print('files merged: %s' % ','.join(file_names))
        if save:
            saved_path = '{}/merged.xlsx'.format(self.home)
            self.renderer.to_excel(df_merged, saved_path)
        return df_merged, file_names

    def to_excel(self, saved_path):
        if self.df is not None:
            self.renderer.to_excel(self.df, saved_path)

    def render_excel(self, df, excel_template, saved_path):
        self.renderer.render_excel(df, saved_path, excel_template)


class ExPlay():
    def __init__(self, home=None, proj_name=None):
        self.home = home if home else os.getcwd()
        self._sources = dict()
        self._parse_yml(proj_name)
        from explay.utils import pd_set_option; pd_set_option(max_colwidth=40, max_columns=15)

    def _parse_yml(self, proj_name=None):
        if proj_name:
            yml_files = ['{}/{}.yml'.format(self.home, proj_name)]
        else:
            yml_files = glob.glob(f'{self.home}/*.yml')
        convs, mergs, parss, rends, projs, outs = [], [], [], [], [], []
        for f in yml_files:

            each = to_yml(f, True)
            converter = each['xlconverter']
            merger = each['xlmerger']
            parser = each['xlparser']
            renderer = each['xlrenderer']
            project = each['xlproject']
            out = each['xloutput']

            if converter: convs.append([f, converter])
            if merger: mergs.append([f, merger])
            if parser: parss.append([f, parser])
            if renderer: rends.append([f, renderer])
            if project: projs.append([f, project])
            if out: outs.append([f, out])

        assert all([len(e)<2 for e in [convs, mergs, parss, rends, projs, outs]])
        self._conv_file = convs[0][0] if convs else None
        self._merg_file = mergs[0][0] if mergs else None
        self._pars_file = parss[0][0] if parss else None
        self._rend_file = rends[0][0] if rends else None
        self._proj_file = projs[0][0] if projs else None
        self._out_file = outs[0][0] if outs else None
        self._conv_params = convs[0][1] if convs else None
        self._merg_params = mergs[0][1] if mergs else None
        self._pars_params = parss[0][1] if parss else None
        self._rend_params = rends[0][1] if rends else None
        self._proj_params = projs[0][1] if projs else None
        self._out_params = outs[0][1] if outs else None

        self._converter = xlConverter(self._conv_params) if convs else None
        self._parsers = [xlBinaryParser(defaultdict(str, each_params)) 
                for each_params in self._pars_params] if parss else []
        self._renderer = xlRenderer(self._rend_params) if rends else None

        self._template = xlTemplate(self._out_params) if outs else None
        if self._proj_params:
            self._project = yaml.dump(self._proj_params, indent=True, default_flow_style=False)
        else:
            self._project = None

    def show_config(self):
        if self._converter:
            print('************************')
            print('*       converter      *')
            print('************************')
            print(self._converter)

        if self._merg_params:
            print('************************')
            print('*         merger       *')
            print('************************')
            merg_print = yaml.dump(self._merg_params, allow_unicode=True, indent=True)
            print(merg_print)

        if self._parsers:
            print('************************')
            print('*         parser       *')
            print('************************')
            for each_parser in self._parsers:
                print(each_parser, '\n')

        if self._renderer:
            print('************************')
            print('*       renderer       *')
            print('************************')
            print(self._renderer)
        
        if self._project:
            print('************************')
            print('*       project        *')
            print('************************')
            print(self._project)
        
    def _merge_sheets(self, conv_name, xlsx_path, sheet_names):
        xlsx_dir = os.path.abspath(os.path.dirname(xlsx_path))
        self.merger = xlMerger(self._conv_params, xlsx_dir)
        df_merged = self.merger.merge_sheets(conv_name, xlsx_path, sheet_names)
        self._sources[conv_name] = df_merged
        return df_merged

    def merge_sheets(self, conv_name, xlsx_path, sheet_names, save=False):
        print('sheets of file %s merged.' % xlsx_path)
        df_merged = self._merge_sheets(conv_name, xlsx_path, sheet_names)
        if save:
            saved_path = '{}/{}_merged.xlsx'.format(self.home, conv_name)
            xlRenderer.to_excel(df_merged, saved_path)
        else:
            print(df_merged)

    def _get_abs_source_path(self, xlsx_dir=None):
        if xlsx_dir:
            if os.path.isabs(xlsx_dir):
                source_path = xlsx_dir
            else:
                source_path = os.path.join(self.home, xlsx_dir)
        else:
            source_path = self.home
        return source_path

    def _merge_files(self, conv_name, relative_paths, xlsx_dir=None, sheet_name=0):
        source_path = self._get_abs_source_path(xlsx_dir)
        filepaths = [os.path.join(source_path, f) for f in relative_paths]
        self.merger = xlMerger(self._conv_params, source_path)
        df_merged = self.merger.merge_files(conv_name, filepaths, sheet_name)
        self._sources[conv_name] = df_merged
        return df_merged

    def merge_files(self, conv_name, relative_paths, xlsx_dir=None, sheet_name=0, save=False):
        source_path = self._get_abs_source_path(xlsx_dir)
        filepaths = [os.path.join(source_path, f) for f in relative_paths]
        df_merged = self._merge_files(conv_name, relative_paths, xlsx_dir, sheet_name)
        print('files merged:\n%s' % '\n'.join(filepaths))
        if save:
            saved_path = '{}/{}_merged.xlsx'.format(self.home, conv_name)
            xlRenderer.to_excel(df_merged, saved_path)
        else:
            print(df_merged)

    def _merge_all(self, conv_name, xlsx_dir=None, sheet_name=0, excludes=None):
        source_path = self._get_abs_source_path(xlsx_dir)
        self.merger = xlMerger(self._conv_params, source_path)
        df_merged, file_names = self.merger.merge_all(conv_name, sheet_name, excludes)
        self._sources[conv_name] = df_merged
        return df_merged, file_names

    def merge_all(self, conv_name, xlsx_dir=None, sheet_name=0, excludes=None, save=False):
        df_merged, file_names = self._merge_all(conv_name, xlsx_dir, sheet_name, excludes)
        print('files merged: %s' % ','.join(file_names))
        if save:
            saved_path = '{}/{}_merged.xlsx'.format(self.home, conv_name)
            xlRenderer.to_excel(df_merged, saved_path)
        else:
            print(df_merged)

    def _df_inputs(self):
        if not self._merg_params: return None
        merged_all = {}
        for each in self._merg_params:
            each = defaultdict(str, each)
            merge_type, name, location = each['type'], each['name'], each['location']
            xlsx_dir, xlsx_path = each['xlsx_dir'], each['xlsx_path']
            print('\n', merge_type, name, location)
            converter_name = each['converter_name']

            print(merge_type)
            if merge_type == 'merge_sheets':
                df_merged = self._merge_sheets(converter_name, xlsx_path, each['sheet_names'])

            elif merge_type == 'merge_all':
                excludes = each['excludes']
                df_merged, _ = self._merge_all(converter_name, xlsx_dir,
                        each['sheet_name'], excludes)

            elif merge_type == 'merge_files':
                df_merged = self._merge_files(converter_name, location,
                        xlsx_dir, each['sheet_name'])
            merged_all[name] = df_merged
        return merged_all

    def _run(self, node):
        local_name = dir(__main__)
        _local = lambda x: getattr(__main__, x)
        if 'name' in node:
            del node['name']
        node_name = list(node.keys())[0]
        parser = getattr(__main__, node_name)
        tp = parser.check_ParserType()
        if tp=='binary_parser':
            node_child = node[node_name]
            left = node_child['left']
            right = node_child['right']
            if type(left)==str and left in local_name:
                left_result = _local(left)
            else:
                left_result = self._run(left)

            if type(right)==str and right in local_name:
                right_result = _local(right)
            else:
                right_result = self._run(right)
            return parser(left_result, right_result)

        elif tp=='unary_parser':
            node_key = list(node.keys())[0]
            node_value = node[node_key]
            if type(node_value)==str and node_value in local_name:
                result = _local(node_value)
            else:
                result = self._run(node_value)
            return parser(result)

    def run_proj(self, to_excel=True):
        components = [self._converter, self._merg_params, self._parsers, self._project]
        if not all(components):
            print('please define all explay components!')
            return

        self.export()
        self.results = {}
        for each_proj in self._proj_params:
            proj_name = each_proj['name']
            self.results[proj_name] = self._run(each_proj)

        if to_excel:
            if self._out_params and self._renderer and self._template:
                self._render_excel()
            else:
                self._to_excel()
        else:
            for name, result in self.results.items():
                print('\nproj result: {} (first 10 rows)'.format(name))
                print(result.head(10))

    def _to_excel(self):
        for proj_name, each_result in self.results.items():
            self._renderer.to_excel(each_result,  'out_{}.xlsx'.format(proj_name))

    def _render_excel(self):
        for e in self._template.output:
            template_name = e['template']
            proj_result = self.results[e['proj_result']]
            path=  e['path']
            self._renderer.render_excel(proj_result, path, template_name)

    def export(self):
        inputs = self._df_inputs()
        for input_name, each_df in inputs.items():
            setattr(__main__, input_name, each_df)
        for each_parser in self._parsers:
            setattr(__main__, each_parser.name, each_parser)


if __name__ == "__main__":
    
    pass
    #ee = ExPlay()
    #ee.export(locals())

    #x1 = ACTION1(df, df_gender)
    #x2 = ACTION2(x1, df_code)
    #DF = ACTION3(ACTION1._output.output[0], x2)
