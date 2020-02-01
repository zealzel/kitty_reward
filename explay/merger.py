# -*- coding: utf-8 -*-

from __future__ import absolute_import
from __future__ import division
from __future__ import print_function

import os
import re
import sys
import json
import yaml
import regex
from copy import copy
import datetime
from collections import namedtuple, defaultdict
import inspect
import functools
import glob
import numpy as np

import xlrd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font

from pandas import DataFrame
import pandas as pd
import pdb

from explay.utils import is_buildin, replace_str
from explay.openpyxl_ext import insert_rows

from explay.utils import register_func
from explay.agg_func import agg_functions
from explay.post_func import common_funcs



def compose(*functions):
    def compose2(f, g):
        return lambda x: f(g(x))
    return functools.reduce(compose2, functions, lambda x: x)


class xlConverter():
    def __init__(self, params):
        self.params = dict([(e['name'], {k:v for k,v in e.items() if k!='name'}) for e in params])

    def __repr__(self):
        msg = '[Converter]\n'
        for name, param in self.params.items():
            msg += ' converter_name: %s\n' % name
            #  for each in param:
                #  msg += str(each)
        return msg

    def _load_excel(self, filepath, sheet_name, first_row, idx_colname, resetindex=True):
        col_indexes, col_names = list(zip(*list(idx_colname.items())))

        # handle column data type
        types = []
        col_names = list(col_names)
        for i, c in enumerate(col_names):
            name_type = c.split('=')
            if len(name_type) > 1: # has type definition
                types.append(name_type[-1].strip())
                col_names[i] = name_type[0].strip()
            else:
                types.append(None)

        parse_cols = [c-1 for c in col_indexes]
        df = pd.read_excel(filepath, sheet_name=sheet_name,
                skiprows=first_row-1, header=None, usecols=parse_cols, names=col_names)
        df = df.reset_index(drop=True) if resetindex else df

        return df, types

    def load_excel(self, converter_name, filepath, sheet_name=0, resetindex=True):
        if type(converter_name)==list:
            output = []
            for each in converter_name:
                first_row = self.params[each]['first_row']
                idx_colname = self.params[each]['idx_colname']
                df, types = self._load_excel(filepath, sheet_name, first_row, idx_colname, resetindex)
                import pdb; pdb.set_trace()
                output.append(df)
        else:
            first_row = self.params[converter_name]['first_row']
            idx_colname = self.params[converter_name]['idx_colname']
            df, types = self._load_excel(filepath, sheet_name, first_row, idx_colname, resetindex)
            output = df

        if 'dropna' in self.params[converter_name]:
            df.dropna(subset=[self.params[converter_name]['dropna']], inplace=True)

        if 'trim' in self.params[converter_name]:
            cols_trim = self.params[converter_name]['trim']
            if type(cols_trim) != list:
                cols_trim = [cols_trim]
            df[cols_trim] = df[cols_trim].applymap(lambda x: str(x).strip())

        for col, col_type in zip(df.columns, types):
            if col_type:
                cast = {'int': compose(int, float), 'float': float}
                df[col] = df[col].apply(cast[col_type])

        return output


class xlMerger(xlConverter):
    def __init__(self, params, source_path):
        xlConverter.__init__(self, params)
        self.source_path = source_path

    def merge_files(self, converter_name, filepaths, sheet_name=0):
        df_all = []
        for filepath in filepaths:
            df_all.append(self.load_excel(converter_name, filepath, sheet_name))
        df = pd.concat(df_all, axis=0)
        df.index = range(len(df))
        return df

    def merge_sheets(self, converter_name, filepath, sheet_names):
        df_list = []
        for sheet_name in sheet_names:
            df_each = self.load_excel(converter_name, filepath, sheet_name)
            df_list.append(df_each)
        df_merged = pd.concat(df_list)
        return df_merged

    def merge_all(self, converter_name, sheet_name=0, filename_excludes=None):
        files = glob.glob('{}/*xlsx*'.format(self.source_path))
        file_names = [os.path.basename(f) for f in files]
        if filename_excludes:
            file_names = list(filter(lambda f: f not in filename_excludes, file_names))
        else:
            file_names = file_names

        df_all = []
        for f in file_names:
            file_path = os.path.join(self.source_path, f)
            df_all.append(self.load_excel(converter_name, file_path, sheet_name))
        df = pd.concat(df_all, axis=0)
        df.index = range(len(df))
        return df, file_names


#  class xlParser():
    #  def __init__(self, jobs):
        #  self.jobs = jobs
        #  self.outputs = defaultdict(list)

    #  def parse(self, df_input, job_name=None):
        #  if job_name in self.outputs:
            #  self.outputs[job_name] = list()

        #  for jobname, operations in self.jobs.items():
            #  if jobname != job_name: continue
            #  df = df_input

            #  for i, each_op in enumerate(operations):
                #  df = each_op.parse(df)
                #  self.outputs[jobname].append(df)

    #  def show_process(self, job_name):
        #  for i, o in enumerate(self.outputs[job_name], 1):
            #  print('\n\noperation %d' % i)
            #  print(o.head())

