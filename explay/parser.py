# -*- coding: utf-8 -*-

from __future__ import absolute_import
from __future__ import division
from __future__ import print_function

import os
import re
import sys
import json
import yaml
import regex
from copy import copy
import datetime
from collections import namedtuple, defaultdict
import inspect
import functools
import glob
import numpy as np

import xlrd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font

from pandas import DataFrame
import pandas as pd
import pdb

from explay.utils import is_buildin, replace_str, register_func
from explay.openpyxl_ext import insert_rows

from explay.agg_func import agg_functions
from explay.post_func import common_funcs



def rpartial(func, *args):
    return lambda *a: func(*(a + args))


def compose(*functions):
    def compose2(f, g):
        return lambda x: f(g(x))
    return functools.reduce(compose2, functions, lambda x: x)


def register_custom_func(name, func):
    global common_funcs
    common_funcs[name] = func


class Operation():
    def __init__(self, params):
        self.params = params
        self.type = params['type']
        self.args = params['args']
        self.name = params['name'] if 'name' in params else ''
        self.load()

    def load(self):
        pass
    
    def __repr__(self):
        operation_type = self.__class__.__name__
        msg = '[%s][%s]' % (operation_type, self.name)
        return msg

        #  return '<{}>'.format(self.__class__.__name__)


class UnaryOperation(Operation):
    def __init__(self, params):
        Operation.__init__(self, params)
        self.op_type = 'unary'


class BinaryOperation(Operation):
    def __init__(self, params):
        Operation.__init__(self, params)
        self.op_type = 'binary'


class Sort(Operation):
    def __init__(self, params):
        UnaryOperation.__init__(self, params)

    def load(self):
        #  print('Sort load')
        self.values = self.args['values']

    def parse(self, df):
        sort_values = self.values
        output = df.sort_values(sort_values)
        return output


class GroupBy(UnaryOperation):
    def __init__(self, params):
        UnaryOperation.__init__(self, params)

    def load(self):
        #  print('Groupby load')
        self.by = self.args['by']
        self.agg = self.args['agg']
        self.eval_func()

    def eval_func(self):
        for name, func_with_arg in self.agg.items():
            title, func_str = func_with_arg[:2]
            if len(func_with_arg) > 1:
                func_args = func_with_arg[2:]

            if not is_buildin(func_str):
                try:
                    agg_func = agg_functions[func_str]
                    if func_args:
                        agg_func = rpartial(agg_func, *func_args)
                except NameError as ex:
                    raise
            else:
                agg_func = eval(func_str)
            self.agg[name] = [title, agg_func]

    def parse(self, df):
        group_by = self.by
        agg_dict = self.agg
        D = df.groupby(group_by, sort=False)

        dataframes = []
        for name, values in agg_dict.items():
            col, func = values
            df_grouped = pd.DataFrame(D[col].apply(lambda x: func(x)))
            df_grouped.columns = [name]
            dataframes.append(df_grouped)
        output = pd.concat(dataframes, axis=1)
        output = output.reset_index(group_by)

        return output


#  class Trim(UnaryOperation):
    #  def __init__(self, params):
        #  UnaryOperation.__init__(self, params)

    #  def load(self):
        #  print('Trim load')
        #  args = self.params['args']
        #  self.columns = args['columns']
        #  self.reset_index = args['reset_index']


    #  def parse(self, df):
        #  output = df.reset_index(df.index.name)[self.columns]
        #  output.index = range(1, len(output)+1)
        #  return output


class Extension(UnaryOperation):
    def __init__(self, params):
        UnaryOperation.__init__(self, params)

    def load(self):
        #  print('Extension load')
        args = self.params['args']
        self.title = args['title']
        self.output_type = args['type']
        self.func = args['func']

    def parse(self, df):
        func_name = self.func

        if self.title in df.columns:
            df_output = df.drop(self.title, axis=1)
        else:
            df_output = df
        output = []

        for index, row in df.iterrows():
            try:
                titles = df.columns.tolist()
                input_dict = dict(zip(titles, row))

                if func_name.startswith('template@'):
                    
                    input_dict.keys()
                    template_string = func_name[9:]
                    p = regex.compile('{.*?\L<options>.*?}', options=input_dict.keys())
                    grouped = [m for m in p.finditer(template_string)]
                    spans = [list(g.span()) for g in grouped]

                    if not grouped:
                        each_output = template_string
                    else:
                        for key, value in input_dict.items():
                            locals()[key] = value

                        for key, value in common_funcs.items():
                            locals()[key] = value

                        values = []
                        for group in grouped:
                            span = group.span()
                            matched = group.group()
                            #  print('matched: %s' % matched)
                            value = eval(matched[1:-1])
                            values.append(value)
                        each_output = replace_str(template_string, spans, values)

                        cast_datetime = rpartial(datetime.datetime.strptime, '%Y-%m-%d %H:%M:%S')

                        if self.output_type:
                            cast = {'int': compose(int, float),
                                    'float': float,
                                    'list': json.loads,
                                    'str': str,
                                    'datetime': cast_datetime}
                            if self.output_type== 'list':
                                each_output = each_output.replace("'", '"')
                            each_output = cast[self.output_type](each_output)
                else:
                    func = global_func[func_name]
                    sig = inspect.signature(func)
                    arg_names = list(sig.parameters)
                    if len(arg_names)==1:
                        inputs = {arg_names[0]: input_dict[input[0]]}
                    else:
                        inputs = {k:v for k,v in input_dict.items() if k in arg_names}
                    each_output = func(**inputs)

                output.append(each_output)

            except AssertionError as ex:
                print('template後的參數未定義')
                import pdb; pdb.set_trace()
                #  raise AssertionError

        df_ext = pd.DataFrame({self.title: output}, index = df.index)
        df_output = pd.concat((df_output, df_ext), axis=1)
        return df_output


class Join(BinaryOperation):
    def __init__(self, params):
        BinaryOperation.__init__(self, params)

    def load(self):
        #  print('Join load')
        args = self.params['args']
        self.on = args['join_on']
        self.how = args['how']
        self.right_cols = args['right_cols']
        self.left_fillna = args['left_fillna']

    def parse(self, left, right):
        if self.right_cols:
            right = right[self.on + self.right_cols]
        joined_df = pd.merge(left, right, on=self.on, how=self.how)

        if self.left_fillna:
            col_to_fillna, replaced = self.left_fillna
            isnull = getattr(joined_df, col_to_fillna).isnull()
            joined_df.loc[isnull, col_to_fillna] = [ [[]] * isnull.sum() ]
        return joined_df


class Condition():
    def __init__(self, variable, operand, value):
        self.var = variable
        self.op = operand
        self.value = value

    def query(self, df):
        if self.op in ['==', 'eq']:
            new_df = df[df.apply(lambda x: x[self.var] == self.value, axis=1)]
        elif self.op == 'in':
            new_df = df[df.apply(lambda x: x[self.var] in self.value, axis=1)]
        elif self.op == 'not_in':
            new_df = df[df.apply(lambda x: x[self.var] not in self.value, axis=1)]

        return new_df


class Filter(UnaryOperation):
    def __init__(self, params):
        UnaryOperation.__init__(self, params)

    def load(self):
        #  print('Filter load')
        args = self.params['args']
        self.title = args['title']
        self.condition = args['condition']

    def parse(self, df):
        if self.condition:
            cond_type = self.condition['type']
            items = self.condition['items']
            conditions = [Condition(*item) for item in items]
            if cond_type == 'once':
                one_cond = conditions[0]
                result = one_cond.query(df)

            elif cond_type == 'and':
                pass

            elif cond_type == 'or':
                pass
        else:
            result = df

        if self.title:
            result = result[self.title]

        return result


class Pivot(UnaryOperation):
    def __init__(self, params):
        UnaryOperation.__init__(self, params)

    def load(self):
        #  print('Pivot load')
        args = self.params['args']
        self.index = args['index']
        self.columns = args['columns']
        self.values = args['values']
        self.reset_index = args['reset_index']

    def parse(self, df):
        index = self.index
        result = df.pivot(index=self.index, columns=self.columns, values=self.values)
        if self.reset_index:
            result = result.reset_index()
        return result


class Melt(UnaryOperation):
    def __init__(self, params):
        UnaryOperation.__init__(self, params)

    def load(self):
        #  print('Melt load')
        args = self.params['args']
        self.id_vars = args['id_vars'] if args['id_vars'] else None
        self.value_vars = args['value_vars'] if args['value_vars'] else None
        self.value_name = args['value_name'] if args['value_name'] else None
        self.var_name = args['var_name'] if args['var_name'] else 'value'
        
    def parse(self, df):
        if 'by_range' in self.value_vars:
            rng = self.value_vars['by_range']
            value_vars = df.columns[rng[0]-1:rng[1]-1]
        elif 'by_name' in self.value_vars:
            pass
        
        result = df.melt(id_vars=self.id_vars,
                 value_vars=value_vars,
                 var_name=self.var_name,
                 value_name=self.value_name)
        return result


def melt_parser(op_params):
    item = Melt(op_params)
    return [item]
        

def pivot_parser(op_params):
    item = Pivot(op_params)
    return [item]
        

def sort_parser(op_params):
    item = Sort(op_params)
    return [item]
        

def filter_parser(op_params):
    item = Filter(op_params)
    return [item]
        

def extend_parser(op_params):
    operations = []
    op_args = op_params['args']
    arg_title = op_args['title']
    arg_title = [arg_title] if type(arg_title)==str else arg_title
    arg_funcs = [e.rstrip() for e in op_args['func'].split('@')]
    len_title, len_func = len(arg_title), len(arg_funcs)-1

    if 'type' not in op_args:
        arg_types = [''] * len_title
    else:
        arg_types = op_args['type']
        arg_types = [arg_types] if type(arg_types)==str else arg_types

    len_type = len(arg_types)

    if len_title != len_func:
        raise Exception('ProcessError',
            'number of extension functions must be pairsed with titles!')

    elif len_title != len_type:
        raise Exception('ProcessError',
            'number of extension types must be pairsed with titles!')

    elif arg_funcs[0] != 'template':
        raise Exception('ProcessError', 'Extension Format not matched!')

    each_op_args = defaultdict(str)

    each_params = op_params.copy()
    for arg_title, arg_func, arg_type in zip(arg_title, arg_funcs[1:], arg_types):
        each_op_args['title'] = arg_title
        each_op_args['func'] = 'template@{}'.format(arg_func)
        each_op_args['type'] = arg_type

        each_params['args'] = each_op_args
        item = Extension(each_params)
        operations.append(item)

    return operations
        

def groupby_parser(op_params):
    item = GroupBy(op_params)
    return [item]


def trim_parser(op_params):
    item = Trim(op_params)
    return [item]


def join_parser(op_params):
    item = Join(op_params)
    return [item]


class xlParser():

    def __init__(self, initializer):
        self._operations = []
        self.output = []
        self._operations = self.parse_schema(initializer)
        #  register_func()

    @classmethod
    def parse_schema(cls, schema):
        _operations = []
        for each_operation in schema:
            op_type = each_operation['type']
            op_args = defaultdict(str, each_operation['args'])
            parsers = {'group_by': groupby_parser,
                       'extend': extend_parser,
                       'trim': trim_parser,
                       'join': join_parser,
                       'filter': filter_parser,
                       'melt': melt_parser,
                       'pivot': pivot_parser,
                       'sort': sort_parser,
                      }
            op_parser = parsers[op_type]
            op_params = each_operation.copy()
            op_params['args'] = op_args
            each_operations = op_parser(op_params) # could have more than one Operation
            _operations.extend(each_operations)
        return _operations

    def __getitem__(self, position):
        return self._operations[position]

    def __len__(self):
        return len(self._operations)

    def __repr__(self):
        return '\n'.join([str(e) for e in self._operations])
        
    def __call__(self, left, right=None):
        return self.parse(left, right)

    def parse(self, left, right=None):
        register_func()
        df = left
        for i, each_op in enumerate(self._operations):
            print(i, each_op)
            if each_op.op_type == 'unary':
                df = each_op.parse(df)
            else:
                df = each_op.parse(df, right)
            self.output.append(df)
        result = self.output[-1]
        return result

    def show_outputs(self, num_for_each=5):
        for i, (each, op) in enumerate(zip(self.output, self._operations)):
            print('[index: %d][%s][%s]' % (i, op, op.name))
            print(each.head(num_for_each), '\n\n')


class xlBinaryParser():

    def __init__(self, binary_initializer):
        self.name = binary_initializer['name']
        left = binary_initializer['left']
        right = binary_initializer['right']
        output = binary_initializer['output']
        self._left = xlParser(left)
        self._right = xlParser(right)
        self._output = xlParser(output)

    def num_of_binary_parser(self, operations):
        op_basetype = lambda op: op.__class__.__bases__[0]
        op_basetypes = [op_basetype(o) for o in operations]
        num = len(list(filter(lambda o: o==BinaryOperation, op_basetypes)))
        return num

    def check_ParserType(self):
        baseclass = self._output[0].__class__.__bases__[0]
        check1 = self._output is not None
        unary_check1 = (not self._left and not self._right)
        unary_check2 = self.num_of_binary_parser(self._output)==0
        binary_check1 = self.num_of_binary_parser(self._output)==1
        binary_check2 = (self.num_of_binary_parser(self._left) + 
                         self.num_of_binary_parser(self._right))==0
        if not check1:
            return None
        elif unary_check1 and unary_check2:
            return 'unary_parser'
        elif binary_check1 and binary_check2:
            return 'binary_parser'

    def __repr__(self):
        left_ops = '\n'.join([' %s' % e for e in self._left])
        right_ops = '\n'.join([' %s' % e for e in self._right])
        output_ops = '\n'.join([' %s' % e for e in self._output])
        left_repr = '[left]\n%s' % left_ops if left_ops else ''
        right_repr = '[right]\n%s' % right_ops if right_ops else ''
        output_repr = '[output]\n%s' % output_ops
        all_repr = [left_repr, right_repr, output_repr]
        all_repr = list(filter(lambda x: x, all_repr))
        all_parsers = '\n'.join(all_repr)

        if self.check_ParserType() == 'unary_parser':
            my_type = '[Unary Parser][%s]' % self.name
        elif self.check_ParserType() == 'binary_parser':
            my_type = '[Binary Parser][%s]' % self.name
        else:
            raise
        return '%s\n%s' % (my_type, all_parsers)

    def __call__(self, left, right=None):
        return self.parse(left, right)

    def parse(self, left, right=None):
        if not self._left:
            print('no left!')
        if not self._right:
            print('no right!')

        if self.check_ParserType() == 'unary_parser':
            output = self._output(left)
        else:
            left_output = self._left.parse(left) if self._left else left
            right_output = self._right.parse(right) if self._right else right
            output = self._output.parse(left_output, right_output)

        return output

    def show_outputs(self, num_for_each=5):
        print('[LEFT]')
        self._left.show_outputs(num_for_each)
        print('[RIGHT]')
        self._right.show_outputs(num_for_each)
        print('[OUTPUT]')
        self._output.show_outputs(num_for_each)
